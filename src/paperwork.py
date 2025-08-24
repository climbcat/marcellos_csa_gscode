
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

evaluator = None



def create_printable_sheets(src_filename, dest_filename):
    print("createPrintableSheets()")

    # load source workbook
    src_wb = load_workbook(src_filename)
    src_sheets = src_wb.worksheets

    # create destination workbook
    dest_wb = Workbook()
    dest_wb.remove(dest_wb.active)  # remove default sheet
    dest_wb.create_sheet("SharedTotals")  # first sheet
    dest_wb.save(dest_filename)

    # create template from first sheet
    create_template(dest_wb, src_sheets[0])
    template_sheet = dest_wb["template"]

    totals_sheet = dest_wb["SharedTotals"]


    for src_sheet in src_sheets:
        print("Processing sheet:", src_sheet.title)
        update_shared_totals_page(totals_sheet, src_sheet)
        #create_group_totals_page(dest_wb, src_sheet)
        #create_member_pages(dest_wb, src_sheet, template_sheet)

    # remove template sheet
    dest_wb.remove(template_sheet)
    dest_wb.save(dest_filename)


def create_template(dest_wb, src_sheet):
    template_sheet = dest_wb.create_sheet("template")

    # delete extra columns beyond 5
    while template_sheet.max_column > 5:
        template_sheet.delete_cols(6)

    # TODO: rewrite using the style of update_shared_totals, which is now the amazing gold standard for doing exactly that

    # copy column A values
    for r, row in enumerate(src_sheet.iter_rows(min_col=1, max_col=1, values_only=True), start=2):
        template_sheet.cell(row=r, column=1, value=row[0])


def update_shared_totals_page(totals_sheet, src_sheet):


    row_idx_src = 6
    col_idx_src = 2

    row_idx_totals = 2
    col_idx_totals = totals_sheet.max_column + 1

    for row in src_sheet.iter_rows(min_row=2, max_col=3, values_only=True):

        cell_totals = totals_sheet.cell(row_idx_totals, col_idx_totals)

        control_string = "%s!%s" % (src_sheet.title, f"{get_column_letter(col_idx_src)}{row_idx_src}")
        value = evaluator.evaluate(control_string).value
        cell_totals.value = value

        row_idx_src += 1
        row_idx_totals += 1


def create_group_totals_page(dest_wb, src_sheet):
    current_sheet_name = f"Total order {src_sheet.title}"
    dest_sheet = dest_wb.copy_worksheet(src_sheet)
    dest_sheet.title = current_sheet_name

    # add group info
    dest_sheet["A1"] = src_sheet["A1"].value

    # add values of total orders (copy col B)
    for r, row in enumerate(src_sheet.iter_rows(min_col=2, max_col=2, values_only=True), start=1):
        dest_sheet.cell(row=r, column=2, value=row[0])

    # insert new row at top
    dest_sheet.insert_rows(1)
    dest_sheet["A1"] = current_sheet_name
    dest_sheet["A1"].font = Font(size=15, bold=True)

    # delete extra columns beyond col 2
    while dest_sheet.max_column > 2:
        dest_sheet.delete_cols(3)


def create_member_pages(dest_wb, src_sheet, template_sheet):
    src_row_count = src_sheet.max_row

    # NOTE value in A1 (contact info)
    note = src_sheet.cell(row=1, column=1).value
    template_sheet.cell(row=2, column=1, value=note)

    number_of_sheets = -(-(src_sheet.max_column - 2) // 4)  # ceil((cols-2)/4)
    for i in range(number_of_sheets):
        current_sheet_name = f"{src_sheet.title} {i+1} of {number_of_sheets}"
        print(current_sheet_name)

        dest_sheet = dest_wb.copy_worksheet(template_sheet)
        dest_sheet.title = current_sheet_name

        # copy 4-column block of data
        start_col = 2 + 4 * i
        for r, row in enumerate(src_sheet.iter_rows(min_col=start_col,
                                                    max_col=start_col+3,
                                                    max_row=src_row_count,
                                                    values_only=True), start=2):
            for j, value in enumerate(row, start=2):
                dest_sheet.cell(row=r, column=j, value=value)

        # set header
        dest_sheet.cell(row=1, column=1, value=current_sheet_name).font = Font(size=15, bold=True)





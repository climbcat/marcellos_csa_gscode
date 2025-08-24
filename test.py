from openpyxl import load_workbook

# Reading an Excel file
wb = load_workbook("testheet.xlsx")
ws = wb.active

# Reading data from a specific cell
data = ws['A1'].value
print("A1 = ", data)

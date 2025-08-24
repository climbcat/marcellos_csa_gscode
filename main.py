#!/usr/bin/env python
# -*- coding: utf-8 -*-
import logging
import argparse
import json
import os

from openpyxl import load_workbook


def main(args):
    logging.basicConfig(format='%(levelname)-5s %(asctime)s: %(message)s', datefmt='%H:%M:%S', level=logging.DEBUG)



    # Reading an Excel file
    print(args.input_file)
    wb = load_workbook(args.input_file)
    ws = wb.active

    # Reading data from a specific cell
    data = ws['A1'].value
    print("A1 = ", data)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('input_file', help='Output file name')
    args = parser.parse_args()

    main(args)
